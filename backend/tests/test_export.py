"""Tests for CSV and XLSX export endpoints."""

from __future__ import annotations

from unittest.mock import patch, MagicMock

import pytest
from httpx import AsyncClient, ASGITransport

from app.main import app
from app.models import (
    AnalysisMetadata,
    AnalysisResult,
    InterestItem,
    KeywordItem,
    VideoItem,
)


@pytest.fixture()
def fake_analysis_result() -> AnalysisResult:
    return AnalysisResult(
        metadata=AnalysisMetadata(
            video_count=2,
            period_days=7,
            region="all",
            run_date="2026-03-20 00:00:00 UTC",
            queries_used=["tech review"],
            errors=[],
        ),
        videos=[
            VideoItem(
                video_id="vid001",
                title="Test Video 1",
                channel="TestChannel",
                views=100000,
                likes=5000,
                published_at="2026-03-15T10:00:00Z",
                language="en",
                thumbnail_url="https://example.com/thumb1.jpg",
            ),
            VideoItem(
                video_id="vid002",
                title="Test Video 2",
                channel="OtherChannel",
                views=50000,
                likes=2500,
                published_at="2026-03-14T08:00:00Z",
                language="ko",
                thumbnail_url="https://example.com/thumb2.jpg",
            ),
        ],
        keywords=[
            KeywordItem(rank=1, keyword="iphone", count=10, category="스마트폰/모바일"),
            KeywordItem(rank=2, keyword="galaxy", count=8, category="스마트폰/모바일"),
            KeywordItem(rank=3, keyword="laptop", count=5, category="노트북/PC"),
        ],
        interests=[
            InterestItem(rank=1, category="스마트폰/모바일", score=18, ratio=1.0),
        ],
    )


# -----------------------------------------------------------------------
# CSV export
# -----------------------------------------------------------------------


class TestExportCSV:
    @pytest.mark.anyio
    async def test_csv_returns_200(self, fake_analysis_result):
        mock_analyzer = MagicMock()
        mock_analyzer.analyze.return_value = fake_analysis_result

        with patch("app.main._get_analyzer", return_value=mock_analyzer):
            transport = ASGITransport(app=app)
            async with AsyncClient(transport=transport, base_url="http://test") as client:
                resp = await client.get("/api/export/csv")

        assert resp.status_code == 200

    @pytest.mark.anyio
    async def test_csv_content_type(self, fake_analysis_result):
        mock_analyzer = MagicMock()
        mock_analyzer.analyze.return_value = fake_analysis_result

        with patch("app.main._get_analyzer", return_value=mock_analyzer):
            transport = ASGITransport(app=app)
            async with AsyncClient(transport=transport, base_url="http://test") as client:
                resp = await client.get("/api/export/csv")

        assert "text/csv" in resp.headers["content-type"]

    @pytest.mark.anyio
    async def test_csv_has_header_row(self, fake_analysis_result):
        mock_analyzer = MagicMock()
        mock_analyzer.analyze.return_value = fake_analysis_result

        with patch("app.main._get_analyzer", return_value=mock_analyzer):
            transport = ASGITransport(app=app)
            async with AsyncClient(transport=transport, base_url="http://test") as client:
                resp = await client.get("/api/export/csv")

        content = resp.text
        lines = content.strip().split("\n")
        assert lines[0].strip() == "rank,keyword,count,category"
        assert len(lines) == 4  # header + 3 keywords

    @pytest.mark.anyio
    async def test_csv_content_disposition(self, fake_analysis_result):
        mock_analyzer = MagicMock()
        mock_analyzer.analyze.return_value = fake_analysis_result

        with patch("app.main._get_analyzer", return_value=mock_analyzer):
            transport = ASGITransport(app=app)
            async with AsyncClient(transport=transport, base_url="http://test") as client:
                resp = await client.get("/api/export/csv")

        assert "attachment" in resp.headers.get("content-disposition", "")
        assert "techpulse_keywords_" in resp.headers.get("content-disposition", "")

    @pytest.mark.anyio
    async def test_csv_keyword_data(self, fake_analysis_result):
        mock_analyzer = MagicMock()
        mock_analyzer.analyze.return_value = fake_analysis_result

        with patch("app.main._get_analyzer", return_value=mock_analyzer):
            transport = ASGITransport(app=app)
            async with AsyncClient(transport=transport, base_url="http://test") as client:
                resp = await client.get("/api/export/csv")

        content = resp.text
        assert "iphone" in content
        assert "galaxy" in content
        assert "laptop" in content


# -----------------------------------------------------------------------
# XLSX export
# -----------------------------------------------------------------------


class TestExportXLSX:
    @pytest.mark.anyio
    async def test_xlsx_returns_200(self, fake_analysis_result):
        mock_analyzer = MagicMock()
        mock_analyzer.analyze.return_value = fake_analysis_result

        with patch("app.main._get_analyzer", return_value=mock_analyzer):
            transport = ASGITransport(app=app)
            async with AsyncClient(transport=transport, base_url="http://test") as client:
                resp = await client.get("/api/export/xlsx")

        assert resp.status_code == 200

    @pytest.mark.anyio
    async def test_xlsx_content_type(self, fake_analysis_result):
        mock_analyzer = MagicMock()
        mock_analyzer.analyze.return_value = fake_analysis_result

        with patch("app.main._get_analyzer", return_value=mock_analyzer):
            transport = ASGITransport(app=app)
            async with AsyncClient(transport=transport, base_url="http://test") as client:
                resp = await client.get("/api/export/xlsx")

        assert "spreadsheetml" in resp.headers["content-type"]

    @pytest.mark.anyio
    async def test_xlsx_content_disposition(self, fake_analysis_result):
        mock_analyzer = MagicMock()
        mock_analyzer.analyze.return_value = fake_analysis_result

        with patch("app.main._get_analyzer", return_value=mock_analyzer):
            transport = ASGITransport(app=app)
            async with AsyncClient(transport=transport, base_url="http://test") as client:
                resp = await client.get("/api/export/xlsx")

        assert "attachment" in resp.headers.get("content-disposition", "")
        assert "techpulse_analysis_" in resp.headers.get("content-disposition", "")

    @pytest.mark.anyio
    async def test_xlsx_is_valid_workbook(self, fake_analysis_result):
        """Verify the response contains a valid XLSX file."""
        from io import BytesIO
        from openpyxl import load_workbook

        mock_analyzer = MagicMock()
        mock_analyzer.analyze.return_value = fake_analysis_result

        with patch("app.main._get_analyzer", return_value=mock_analyzer):
            transport = ASGITransport(app=app)
            async with AsyncClient(transport=transport, base_url="http://test") as client:
                resp = await client.get("/api/export/xlsx")

        wb = load_workbook(BytesIO(resp.content))
        assert "Keywords" in wb.sheetnames
        assert "Top Videos" in wb.sheetnames

        ws_kw = wb["Keywords"]
        # Header row + 3 keyword rows
        assert ws_kw.max_row == 4

        ws_vid = wb["Top Videos"]
        # Header row + 2 video rows
        assert ws_vid.max_row == 3
